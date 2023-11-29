from flask import Flask, render_template, request, redirect
import openpyxl
from datetime import datetime

app = Flask(__name__)

@app.route('/')
def accueil():
    produits = obtenir_derniers_produits()
    return render_template('accueil.html', produits=produits)

@app.route('/ajouter_produit', methods=['GET', 'POST'])
def ajouter_produit():
    if request.method == 'POST':
        ajouter_produit_formulaire(request.form)
        return redirect('/')  # Redirige vers la page d'accueil après l'ajout

    return render_template('ajouter_produit.html')

def obtenir_derniers_produits():
    classeur_excel = openpyxl.load_workbook('classeur.xlsx')
    ws_donnees = classeur_excel['Donnees']

    produits = []
    for ligne in ws_donnees.iter_rows(min_row=2, values_only=True, max_col=4):
        nom_produit, prix_produit, quantite_produit, date_produit = ligne
        produit = {
            'nom': nom_produit,
            'prix': prix_produit,
            'quantite': quantite_produit,
            'date': date_produit.strftime('%d/%m/%Y')  # Format jour/mois/année
        }
        produits.append(produit)

    # Tri des produits par date (le plus récent en premier)
    produits.sort(key=lambda x: x['date'], reverse=True)

    return produits


def ajouter_produit_formulaire(formulaire):
    classeur_excel = openpyxl.load_workbook('classeur.xlsx')
    ws_donnees = classeur_excel['Donnees']

    nom_produit = formulaire['nom_produit']
    prix_produit = float(formulaire['prix_produit'])
    quantite_produit = float(formulaire['quantite_produit'])
    date_produit = datetime.now()

    derniere_ligne = ws_donnees.max_row + 1
    ws_donnees.cell(row=derniere_ligne, column=1, value=nom_produit)
    ws_donnees.cell(row=derniere_ligne, column=2, value=prix_produit)
    ws_donnees.cell(row=derniere_ligne, column=3, value=quantite_produit)
    ws_donnees.cell(row=derniere_ligne, column=4, value=date_produit)

    classeur_excel.save('classeur.xlsx')

if __name__ == '__main__':
    app.run(debug=True)
